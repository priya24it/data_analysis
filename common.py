import pandas as pd 
import openpyxl
import openpyxl.styles as sty
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment,Font,colors  
import string



#Create the instance of excel workbook.
wb = openpyxl.Workbook()
ws = wb.active

def new_():
  pass

def read_csv(test_data_path,filename):
  sourcedata= pd.read_csv(f'{test_data_path}/{filename}',low_memory=False,encoding='unicode_escape')
  return sourcedata
  
def read_text(test_data_path,filename):
  sourcedata= pd.read_csv(f'{test_data_path}/{filename}',low_memory=False,delimiter='|',encoding='unicode_escape')
  return sourcedata

def read_excel(test_data_path,filename):
  sourcedata= pd.read_excel(f'{test_data_path}/{filename}',low_memory=False,encoding='unicode_escape')
  return sourcedata
  
def get_dataframe_info(df):
  df_types = pd.DataFrame(df.dtypes)
  df_nulls = df.count()
  df_null_count = pd.concat([df_nulls,df_types], axis=1)
  df_null_count = df_null_count.reset_index()
  # Reassign column names
  col_names = ["column_names","non_null_counts", "python_data_types"]
  df_null_count.columns = col_names
  return df_null_count
  
def change_datatype(input_dataframe):
  df1 = input_dataframe.astype(str)
  return df1
  
def cal_of_unique_values(sourcedata,col_name):
  sourcedata[col_name] = sourcedata[col_name].fillna("NULL")
  value_counts = sourcedata[col_name].value_counts()
  df = pd.DataFrame(value_counts)
  df = df.reset_index()
  df.columns = [col_name,"Count"]
  return df

def get_the_first_three_rows(data_frame,column_names):
  data_frame = data_frame[column_names]
  data_frame = data_frame[:3]
  return data_frame
  
def column_names():
  columns = list(string.ascii_uppercase)
  columns_AA = list(map(lambda x:'A'+x,columns))
  columns_BB = list(map(lambda x:'B'+x,columns))
  columns_CC = list(map(lambda x:'C'+x,columns))
  columns_DD = list(map(lambda x:'D'+x,columns))
  columns_EE = list(map(lambda x:'E'+x,columns))
  columns_FF = list(map(lambda x:'F'+x,columns))
  columns = columns + columns_AA + columns_BB + columns_CC + columns_DD + columns_EE + columns_FF
  
  return columns 
  
def append_data_to_excel(data = None,work_sheet = ws,dc = None,dk = None,ind=None):
  cols = column_names() 
  col_name = [] 
  if ind == None and data.shape[0] > 0:
    for col in range(0,data.shape[1]):
      col_name.append(cols[col]+str(work_sheet.max_row+1))  
    work_sheet.cell(row=work_sheet.max_row+1,column=work_sheet.max_column+1).value = ''
    rows = dataframe_to_rows(data, index=False) 
    for r_idx, row in enumerate(rows, work_sheet.max_row):
      for c_idx, value in enumerate(row,1):
        work_sheet.cell(row=r_idx, column=c_idx, value=value) 
        work_sheet.cell(row=r_idx, column=c_idx).font = sty.Font(name='Times New Roman', size=11)
    work_sheet.cell(row=work_sheet.max_row+1,column=work_sheet.max_column+1).value = ''
  return col_name
  
def format_excel(ws):
  columns = column_names()  
  for i in range(0, len(columns)):
    if columns[i] == "A":
      ws.column_dimensions[columns[i]].width = 35
    elif columns[i] == "B":
      ws.column_dimensions[columns[i]].width = 35
    elif columns[i] == "D":
      ws.column_dimensions[columns[i]].width = 35
    elif columns[i] == "E":
      ws.column_dimensions[columns[i]].width = 35
    else:
      ws.column_dimensions[columns[i]].width = 30  

def add_background_color(ws,list_of_names):
  for col in range(0,len(list_of_names)):
    try:
      ws[list_of_names[col]].font = sty.Font(bold=True)
      ws[list_of_names[col]].fill = sty.PatternFill(start_color='00339966', end_color='00339966',fill_type = "solid") 
    except:
      pass #FF0000:red
      
  #Apply the colors for the 2nd row in the sheet i.e df_info dataframe.
  list_of_names_1 = ['A2','B2','C2']
  for col in range(0,len(list_of_names)):
    try:
      ws[list_of_names_1[col]].font = sty.Font(bold=True)
      ws[list_of_names_1[col]].fill = sty.PatternFill(start_color='00339966', end_color='00339966',fill_type = "solid")     
    except:
      pass 

def change_font_color_to_red(ws,c_names):
  for c in range(0,len(c_names)):
    ws[c_names[c]].font = Font(color='FF0F0F',name='Times New Roman', size=11)
    ws[c_names[c]].alignment = Alignment(wrapText=True)

  


